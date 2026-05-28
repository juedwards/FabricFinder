"""Tenant detail lookup, backed by the LXP / FY27 Priority workbook.

The workbook is ``memory/xls_data/EDU Tenant Detail from LXP with Minecraft MAU
- US.xlsx``. Sheet name carries a refresh date (e.g. ``LXP US Tenants
05.27.2026``), so we resolve it by prefix.

The ``LXP US Tenants`` sheet's first row contains section banners (which LXP
table each block of columns came from); the second row is the real header.
"""
from __future__ import annotations

import os
from typing import Optional

import openpyxl

XLSX = os.path.join(
    os.path.dirname(__file__),
    "memory", "xls_data",
    "EDU Tenant Detail from LXP with Minecraft MAU - US.xlsx",
)
SHEET_PREFIX = "LXP US Tenants"

# Subset of columns surfaced to the LLM tool / CSV / PDF. Keep this list short
# so the model gets signal without drowning in 80 columns.
_CONTACT_FIELDS = [
    "Tenant ID", "Tenant Name", "TPID", "TPName",
    "AM", "1) Account Executive",
    "Geo Hierarchy - Country", "Tenant's Region",
    "Field Area", "Field Region", "Vertical",
    "Sales Territory", "Sales Unit",
    "Current Dominant Position", "Cohort", "SKU Group", "Primary Upsell",
    "Primary Expiration Month", "Earliest Expiration Date",
    "Earliest Anniversary Date", "Renewal/Midterm flag", "2) Renewal Date",
    "3) Advisory Board", "7) FY26 Engagement",
    "Date of Engagement ", "Results of FY26 Engagement",
    "Training Partner", "Technical Blocker", "Notes",
    "Included in MCEDU Priority Tenant List?",
    "Current MAU", "MAU YoY", "Current NUA", "NUA YoY",
    "6) Last 12 Months Peak MAU",
    "Minecraft MAU Faculty", "Minecraft MAU Student",
]

# Pretty labels for the small "contact" view returned by contact_summary().
# Order matters: this is the order the bot/PDF will render them.
_CONTACT_LABELS = [
    ("AM", "Account Manager (alias)"),
    ("1) Account Executive", "Account Executive"),
    ("Vertical", "Vertical"),
    ("Field Area", "Field Area"),
    ("Field Region", "Field Region"),
    ("Sales Unit", "Sales Unit"),
    ("Current Dominant Position", "Dominant SKU"),
    ("Primary Upsell", "Primary Upsell"),
    ("2) Renewal Date", "Renewal Date"),
    ("Renewal/Midterm flag", "Renewal/Midterm Flag"),
    ("Earliest Anniversary Date", "Earliest Anniversary"),
    ("3) Advisory Board", "Advisory Board"),
    ("7) FY26 Engagement", "FY26 Engagement"),
    ("Training Partner", "Training Partner"),
    ("Technical Blocker", "Technical Blocker"),
]

_rows: list[dict] = []          # full per-tenant detail dicts
_by_id: dict[str, dict] = {}    # lowercased GUID -> row
_loaded = False


def _resolve_sheet(wb):
    for name in wb.sheetnames:
        if name.startswith(SHEET_PREFIX):
            return wb[name]
    raise RuntimeError(
        f"No sheet starting with {SHEET_PREFIX!r} in {os.path.basename(XLSX)}"
    )


def _load():
    global _loaded
    if _loaded:
        return
    if not os.path.exists(XLSX):
        _loaded = True
        return
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = _resolve_sheet(wb)

    it = ws.iter_rows(values_only=True)
    next(it)                # skip banner row
    header = next(it)
    # Map every kept column to its index (None if absent in this refresh).
    indices = {col: (header.index(col) if col in header else None)
               for col in _CONTACT_FIELDS}
    tid_idx = indices["Tenant ID"]

    for r in it:
        if tid_idx is None or r[tid_idx] is None:
            continue
        tid = str(r[tid_idx]).strip().lower()
        if not tid or tid in _by_id:    # first-wins on duplicates
            continue
        row = {col: (r[i] if i is not None else None)
               for col, i in indices.items()}
        _rows.append(row)
        _by_id[tid] = row
    _loaded = True


def _name_of(row: dict) -> str:
    return str(row.get("Tenant Name") or "").strip()


def details_for_id(tenant_id: str) -> Optional[dict]:
    """Exact GUID lookup. Returns the full curated row dict, or None."""
    _load()
    if not tenant_id:
        return None
    return _by_id.get(str(tenant_id).strip().lower())


def lookup(query: str, limit: int = 10) -> list[dict]:
    """Resolve a query (GUID or name substring) to one or more tenant rows.

    Returns a list ordered: exact name, prefix, then other substring matches.
    """
    _load()
    if not query:
        return []
    q = query.strip().lower()

    # GUID-shaped → direct lookup.
    if len(q) == 36 and q.count("-") == 4:
        row = _by_id.get(q)
        return [row] if row else []

    hits = [r for r in _rows if q in _name_of(r).lower()]
    hits.sort(key=lambda r: (
        _name_of(r).lower() != q,
        not _name_of(r).lower().startswith(q),
        len(_name_of(r)),
    ))
    return hits[:limit]


def contact_summary(tenant_id: str) -> Optional[dict]:
    """Compact contact block for the tenant PDF.

    Returns ``{"items": [(label, value), ...], "tenant_name": ...}``,
    skipping fields that are blank. Returns None if the tenant isn't found.
    """
    row = details_for_id(tenant_id)
    if row is None:
        return None
    items = []
    for col, label in _CONTACT_LABELS:
        v = row.get(col)
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        items.append((label, v))
    return {"tenant_name": _name_of(row), "items": items}


def enrich_columns() -> list[str]:
    """Columns appended to CSV exports that contain a TenantId column."""
    return ["AM", "Account Executive", "Vertical", "Field Region",
            "Renewal Date"]


def enrich_row(tenant_id: str) -> list:
    """Return the enrichment values aligned with :func:`enrich_columns`."""
    row = details_for_id(tenant_id)
    if row is None:
        return [None] * 5
    return [
        row.get("AM"),
        row.get("1) Account Executive"),
        row.get("Vertical"),
        row.get("Field Region"),
        row.get("2) Renewal Date"),
    ]


def count() -> int:
    _load()
    return len(_rows)


if __name__ == "__main__":
    print(f"{count()} tenants loaded from {os.path.basename(XLSX)}")
    for r in lookup("nyc public", limit=3):
        print(f"  {_name_of(r):40} {r.get('Tenant ID')}  AM={r.get('AM')}  "
              f"AE={r.get('1) Account Executive')}")
